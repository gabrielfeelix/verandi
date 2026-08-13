"""Paleta e estilos comuns às planilhas da Verandi."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TINTA   = '1B2430'   # texto principal
GRAFITE = '5A6478'   # texto secundário
LINHA   = 'DDE3EC'
FUNDO   = 'F7F9FC'
MARCA   = '0F5C6E'   # verde-petróleo: cor da Verandi
MARCA_C = 'E6F2F4'
OK      = '1B7F5A'
OK_C    = 'E4F4EC'
ALERTA  = 'B4791A'
ALERTA_C= 'FCF2DF'
CHEIO   = 'F2F5F9'

def fonte(tam=11, negrito=False, cor=TINTA, italico=False):
    return Font(name='Aptos Narrow', size=tam, bold=negrito, color=cor, italic=italico)

def fundo(cor):
    return PatternFill('solid', fgColor=cor)

def borda(cor=LINHA, baixo=True, direita=False):
    lado = Side(style='thin', color=cor)
    return Border(bottom=lado if baixo else None, right=lado if direita else None)

CENTRO = Alignment(horizontal='center', vertical='center')
ESQ    = Alignment(horizontal='left', vertical='center')
QUEBRA = Alignment(horizontal='left', vertical='top', wrap_text=True)

def titulo(ws, celula, texto, tam=16):
    c = ws[celula]; c.value = texto
    c.font = fonte(tam, True, MARCA); c.alignment = ESQ
    return c

def subtitulo(ws, celula, texto):
    c = ws[celula]; c.value = texto
    c.font = fonte(10, False, GRAFITE); c.alignment = QUEBRA
    return c

def cabecalho(ws, linha, colunas, largura=None, cor=MARCA):
    """Escreve a linha de cabeçalho de uma tabela e devolve a próxima linha."""
    for i, nome in enumerate(colunas, start=1):
        c = ws.cell(linha, i, nome)
        c.font = fonte(9, True, 'FFFFFF')
        c.fill = fundo(cor)
        c.alignment = CENTRO
        c.border = borda('FFFFFF', True, True)
    ws.row_dimensions[linha].height = 26
    if largura:
        for i, w in enumerate(largura, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return linha + 1

def zebra(ws, primeira, ultima, ncols):
    for r in range(primeira, ultima + 1):
        if (r - primeira) % 2 == 1:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = fundo(FUNDO)

def moldura(ws):
    ws.sheet_view.showGridLines = False
