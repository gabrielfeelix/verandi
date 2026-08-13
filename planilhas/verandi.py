# -*- coding: utf-8 -*-
"""Verandi — modelo de agenda anual.

A ideia que sustenta o arquivo: **a Grade é a chave mestre.** Ela diz quais
horários existem, com quem e para quantos. Todo o resto — quem está matriculado,
quantas vagas sobram, o que o robô responde — sai dela por fórmula.

Uma linha por horário, e não blocos desenhados. É isso que permite acrescentar
professor, mudar capacidade e abrir horário novo sem redesenhar a planilha.
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.properties import PageSetupProperties
from estilo import *
from openpyxl.styles import Alignment

DIAS = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
SEM_ACENTO = ['segunda', 'terca', 'quarta', 'quinta', 'sexta', 'sabado', 'domingo']
MESES = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
         'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

LIN_GRADE = 140          # linhas reservadas na Grade
LIN_ALUNOS = 400
LIN_MES = 220
G1, G2 = 5, 5 + LIN_GRADE - 1        # primeira e última linha de dados da Grade
A1_, A2 = 5, 5 + LIN_ALUNOS - 1

def caixa(ws, cel_ini, cel_fim, texto, cor=MARCA_C):
    ws.merge_cells(f'{cel_ini}:{cel_fim}')
    c = ws[cel_ini]; c.value = texto
    c.fill = fundo(cor); c.alignment = QUEBRA; c.font = fonte(10, False, TINTA)

def impressao(ws, paisagem=True, repetir=None):
    ws.page_setup.orientation = 'landscape' if paisagem else 'portrait'
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    if repetir: ws.print_title_rows = repetir

# =============================================================== construção
def monta(saida):
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------- Início
    ws = wb.active; ws.title = 'Início'; moldura(ws); impressao(ws, True)
    ws.column_dimensions['A'].width = 3
    for col, w in zip('BCDEFGH', [17, 17, 17, 17, 17, 17, 17]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('B2:B4')
    c = ws['B2']; c.value = 'sua\nlogo'
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.font = fonte(11, True, GRAFITE); c.fill = fundo(MARCA_C)
    ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 22; ws.row_dimensions[4].height = 22

    d = ws['C2']; d.value = '=IF(Config!$C$4="","Seu negócio",Config!$C$4)'
    d.font = fonte(20, True, TINTA); d.alignment = ESQ; ws.merge_cells('C2:E2')
    e = ws['C3']; e.value = '=CONCATENATE("Agenda ",Config!$C$5)'
    e.font = fonte(12, False, GRAFITE); e.alignment = ESQ
    f = ws['C4']; f.value = 'Verandi'; f.font = fonte(10, True, MARCA); f.alignment = ESQ

    cartoes = [
        ('Horários por semana', f'=COUNTA(Grade!$B${G1}:$B${G2})'),
        ('Vagas', f'=SUM(Grade!$E${G1}:$E${G2})'),
        ('Matriculados', f'=SUM(Grade!$F${G1}:$F${G2})'),
        ('Vagas livres', f'=SUM(Grade!$G${G1}:$G${G2})'),
        ('Ocupação', f'=IFERROR(SUM(Grade!$F${G1}:$F${G2})/SUM(Grade!$E${G1}:$E${G2}),0)'),
        ('Alunos ativos', f'=COUNTIF(Alunos!$J${A1_}:$J${A2},"Sim")'),
    ]
    for i, (rot, form) in enumerate(cartoes):
        col = 2 + i
        c = ws.cell(7, col, form); c.font = fonte(20, True, MARCA); c.alignment = CENTRO
        if rot == 'Ocupação': c.number_format = '0%'
        r = ws.cell(8, col, rot); r.font = fonte(9, False, GRAFITE); r.alignment = CENTRO
        for lin in (7, 8): ws.cell(lin, col).fill = fundo(MARCA_C)
    ws.row_dimensions[7].height = 34; ws.row_dimensions[8].height = 20

    passos = [
        ('1 · Config', 'Escreva o nome do negócio e o ano. Cadastre quem atende e o que vocês oferecem. É rápido e só se faz uma vez.'),
        ('2 · Grade', 'Uma linha por horário fixo da semana: dia, hora, quem atende, quantas vagas. Esta aba é o coração — tudo se calcula a partir dela.'),
        ('3 · Alunos', 'Quem é atendido e em qual horário fixo. As vagas livres se atualizam sozinhas.'),
        ('4 · Meses', 'Uma aba por mês, para marcar presença. As datas de cada dia da semana aparecem no topo, já calculadas.'),
        ('5 · AutoFluxos', 'Saída lida pelo robô do WhatsApp. Não precisa mexer.'),
    ]
    linha = 11
    t = ws.cell(10, 2, 'Como usar'); t.font = fonte(13, True, TINTA)
    for tit, txt in passos:
        c = ws.cell(linha, 2, tit); c.font = fonte(10, True, MARCA); c.alignment = ESQ
        d = ws.cell(linha, 3, txt); d.font = fonte(10, False, GRAFITE); d.alignment = QUEBRA
        ws.merge_cells(start_row=linha, start_column=3, end_row=linha, end_column=8)
        ws.row_dimensions[linha].height = 30
        linha += 1
    ws.print_area = f'A1:H{linha + 2}'
    caixa(ws, f'B{linha+1}', f'H{linha+1}',
          'Regra de ouro: nunca apague uma coluna nem mude a ordem das abas. '
          'Acrescentar linha pode sempre; a planilha foi feita para crescer para baixo.')
    ws.row_dimensions[linha+1].height = 34

    # ------------------------------------------------------------- Config
    wc = wb.create_sheet('Config'); moldura(wc); impressao(wc, False)
    for col, w in zip('ABCDEF', [3, 22, 26, 20, 26, 22]): wc.column_dimensions[col].width = w
    titulo(wc, 'B2', 'Configuração')
    subtitulo(wc, 'B3', 'Preencha uma vez. As listas daqui viram as opções das outras abas.')
    wc.merge_cells('B3:F3')

    wc['B4'].value = 'Nome do negócio'; wc['B4'].font = fonte(10, True)
    wc['C4'].fill = fundo(MARCA_C); wc['C4'].font = fonte(11)
    wc['B5'].value = 'Ano'; wc['B5'].font = fonte(10, True)
    wc['C5'].value = 2026; wc['C5'].fill = fundo(MARCA_C); wc['C5'].font = fonte(11); wc['C5'].alignment = ESQ

    p = cabecalho(wc, 8, ['', 'Quem atende', 'Apelido na agenda', 'Ativo'], None)
    for i in range(p, p + 20):
        for c in range(2, 5):
            wc.cell(i, c).border = borda(); wc.cell(i, c).font = fonte(10)
    wc.cell(7, 2, 'Profissionais').font = fonte(12, True, TINTA)

    q = cabecalho(wc, 31, ['', 'O que vocês oferecem', 'Duração (min)', 'Vagas padrão'], None)
    for i in range(q, q + 15):
        for c in range(2, 5):
            wc.cell(i, c).border = borda(); wc.cell(i, c).font = fonte(10)
    wc.cell(30, 2, 'Serviços').font = fonte(12, True, TINTA)

    wc.cell(48, 2, 'Marcas de presença').font = fonte(12, True, TINTA)
    subtitulo(wc, 'B49', 'Use sempre estes códigos. Escrever à mão vira dado que ninguém consegue somar depois.')
    wc.merge_cells('B49:F49'); wc.row_dimensions[49].height = 28
    m = cabecalho(wc, 50, ['', 'Código', 'Significado', 'Conta como falta?'], None)
    marcas = [('P', 'Presente', 'Não'), ('F', 'Faltou', 'Sim'),
              ('FJ', 'Faltou e avisou', 'Sim'), ('REP', 'Veio repor outra aula', 'Não'),
              ('TR', 'Trocou de horário nesta semana', 'Não'),
              ('LIC', 'Afastado por um período', 'Não'),
              ('—', 'Não houve aula (feriado, folga)', 'Não')]
    for i, (cod, sig, falta) in enumerate(marcas):
        r = m + i
        wc.cell(r, 2, cod).font = fonte(10, True); wc.cell(r, 2).alignment = CENTRO
        wc.cell(r, 3, sig).font = fonte(10)
        wc.cell(r, 4, falta).font = fonte(10); wc.cell(r, 4).alignment = CENTRO
        for c in range(2, 5): wc.cell(r, c).border = borda()
    zebra(wc, m, m + len(marcas) - 1, 4)

    wb.defined_names.add(DefinedName('profissionais', attr_text=f'Config!$C${p}:$C${p+19}'))
    wb.defined_names.add(DefinedName('servicos', attr_text=f'Config!$C${q}:$C${q+14}'))
    wc.print_area = f'A1:F{m + 8}'
    wb.defined_names.add(DefinedName('marcas', attr_text=f'Config!$B${m}:$B${m+len(marcas)-1}'))

    # -------------------------------------------------------------- Grade
    wg = wb.create_sheet('Grade'); moldura(wg); impressao(wg, True, '1:4')
    titulo(wg, 'A1', 'Grade da semana')
    subtitulo(wg, 'A2', 'Uma linha por horário fixo. É daqui que sai tudo: vagas livres, ocupação e o que o robô responde.')
    wg.merge_cells('A2:H2'); wg.row_dimensions[2].height = 26
    cols = ['Dia', 'Hora', 'Quem atende', 'Serviço', 'Vagas', 'Matriculados', 'Livres', 'Situação'] + DIAS
    cabecalho(wg, 4, cols, [13, 9, 20, 22, 9, 14, 9, 14])
    for r in range(G1, G2 + 1):
        wg.cell(r, 6, f'=IF($A{r}="","",COUNTIFS(Alunos!$G:$G,$A{r}&" "&$B{r},Alunos!$J:$J,"Sim"))')
        wg.cell(r, 7, f'=IF($A{r}="","",MAX(0,$E{r}-$F{r}))')
        wg.cell(r, 8, f'=IF($A{r}="","",IF($G{r}=0,"Cheio","Tem vaga"))')
        # Uma coluna auxiliar por dia da semana. Existe porque a Grade não é
        # ordenada — quem preenche escreve na ordem que quiser — então não dá
        # para somar faixa contígua. Ficam escondidas.
        for j, dia in enumerate(DIAS):
            col = get_column_letter(9 + j)
            ant = f'${col}{r-1}' if r > G1 else '""'
            wg.cell(r, 9 + j, f'={ant}&IF(AND($A{r}="{dia}",$G{r}>0),$B{r}&";","")')
        for c in range(1, 9):
            cel = wg.cell(r, c); cel.font = fonte(10); cel.border = borda()
            cel.alignment = ESQ if c in (1, 3, 4) else CENTRO
    zebra(wg, G1, G2, 8)
    for j in range(len(DIAS)):
        wg.column_dimensions[get_column_letter(9 + j)].hidden = True
    wg.freeze_panes = f'A{G1}'
    wg.auto_filter.ref = f'A4:H{G2}'
    # Área de impressão limitada. Sem isto, as linhas reservadas em branco
    # viram páginas vazias — o arquivo inteiro saía com 100 páginas. Quem
    # precisar de mais é só esticar a área.
    wg.print_area = f'A1:H{G1 + 55}'
    dv_dia = DataValidation(type='list', formula1=f'="{",".join(DIAS)}"', allow_blank=True)
    wg.add_data_validation(dv_dia); dv_dia.add(f'A{G1}:A{G2}')
    dv_prof = DataValidation(type='list', formula1='=profissionais', allow_blank=True)
    wg.add_data_validation(dv_prof); dv_prof.add(f'C{G1}:C{G2}')
    dv_serv = DataValidation(type='list', formula1='=servicos', allow_blank=True)
    wg.add_data_validation(dv_serv); dv_serv.add(f'D{G1}:D{G2}')
    wg.conditional_formatting.add(f'H{G1}:H{G2}',
        CellIsRule(operator='equal', formula=['"Tem vaga"'], fill=fundo(OK_C), font=fonte(10, True, OK)))
    wg.conditional_formatting.add(f'H{G1}:H{G2}',
        CellIsRule(operator='equal', formula=['"Cheio"'], fill=fundo(CHEIO), font=fonte(10, False, GRAFITE)))

    # ------------------------------------------------------------- Alunos
    wa = wb.create_sheet('Alunos'); moldura(wa); impressao(wa, True, '1:4')
    titulo(wa, 'A1', 'Alunos')
    subtitulo(wa, 'A2', 'Quem é atendido, e em qual horário fixo. O campo Turma usa a lista da Grade.')
    wa.merge_cells('A2:J2'); wa.row_dimensions[2].height = 26
    cols = ['Matrícula', 'Nome', 'Telefone', 'Plano', 'Venc. do plano', 'Início', 'Turma',
            'Próxima avaliação', 'Observações', 'Ativo']
    cabecalho(wa, 4, cols, [11, 32, 16, 16, 14, 12, 20, 16, 30, 8])
    for r in range(A1_, A2 + 1):
        for c in range(1, 11):
            cel = wa.cell(r, c); cel.font = fonte(10); cel.border = borda()
            cel.alignment = ESQ if c in (2, 4, 7, 9) else CENTRO
        for c in (5, 6, 8): wa.cell(r, c).number_format = 'dd/mm/yyyy'
    zebra(wa, A1_, A2, 10)
    wa.freeze_panes = f'C{A1_}'
    wa.auto_filter.ref = f'A4:J{A2}'
    wa.print_area = f'A1:J{A1_ + 55}'
    dv_sn = DataValidation(type='list', formula1='"Sim,Não"', allow_blank=True)
    wa.add_data_validation(dv_sn); dv_sn.add(f'J{A1_}:J{A2}')
    wa.conditional_formatting.add(f'E{A1_}:E{A2}',
        FormulaRule(formula=[f'AND($E{A1_}<>"",$E{A1_}<TODAY())'], fill=fundo(ALERTA_C), font=fonte(10, True, ALERTA)))
    wb.defined_names.add(DefinedName('turmas', attr_text=f'Grade!$A${G1}:$A${G2}'))

    # -------------------------------------------------------------- Meses
    for i, mes in enumerate(MESES, start=1):
        wm = wb.create_sheet(f'{i:02d} {mes[:3]}')
        moldura(wm); impressao(wm, True, '1:8')
        titulo(wm, 'A1', f'=CONCATENATE("{mes} ",Config!$C$5)')
        subtitulo(wm, 'A2', 'Marque a presença usando os códigos da aba Config. As datas de cada dia estão na tabela ao lado.')
        wm.merge_cells('A2:F2'); wm.row_dimensions[2].height = 26

        # datas de cada dia da semana no mês, por fórmula
        rot = wm.cell(4, 12, 'Datas do mês')
        rot.font = fonte(9, True, 'FFFFFF'); rot.fill = fundo(MARCA); rot.alignment = CENTRO
        for c in range(13, 18):
            h = wm.cell(4, c, f'{c-12}ª'); h.font = fonte(9, True, 'FFFFFF')
            h.fill = fundo(MARCA); h.alignment = CENTRO
        for j, dia in enumerate(DIAS):
            r = 5 + j
            wm.cell(r, 12, dia).font = fonte(9, False, GRAFITE)
            for k in range(5):
                base = (f'DATE(Config!$C$5,{i},1)+MOD({j+1}-WEEKDAY(DATE(Config!$C$5,{i},1),2)+7,7)+7*{k}')
                f = f'=IF(MONTH({base})={i},{base},"")'
                c = wm.cell(r, 13 + k, f)
                c.number_format = 'dd/mm'; c.font = fonte(9); c.alignment = CENTRO
                c.border = borda()
        for col, w in zip('LMNOPQ', [14, 9, 9, 9, 9, 9]): wm.column_dimensions[col].width = w
        wm.column_dimensions['K'].width = 3
        # Listra só nas colunas do bloco. Antes ela varria a largura inteira e
        # o bloco de datas parecia sujeira solta no meio da folha.
        for r2 in range(5, 12):
            if (r2 - 5) % 2 == 1:
                for c in range(12, 18): wm.cell(r2, c).fill = fundo(FUNDO)
        for c in range(12, 18): wm.cell(11, c).border = borda()

        cols = ['Dia', 'Hora', 'Quem atende', 'Aluno', '1ª', '2ª', '3ª', '4ª', '5ª', 'Faltas']
        cabecalho(wm, 13, cols, [13, 9, 18, 30, 7, 7, 7, 7, 7, 9])
        prim = 14; ult = prim + LIN_MES - 1
        for r in range(prim, ult + 1):
            for c in range(1, 11):
                cel = wm.cell(r, c); cel.font = fonte(10); cel.border = borda()
                cel.alignment = ESQ if c in (1, 3, 4) else CENTRO
            wm.cell(r, 10, f'=IF($D{r}="","",COUNTIF($E{r}:$I{r},"F")+COUNTIF($E{r}:$I{r},"FJ"))')
        zebra(wm, prim, ult, 10)
        wm.freeze_panes = f'A{prim}'
        wm.auto_filter.ref = f'A13:J{ult}'
        wm.print_area = f'A1:Q{prim + 55}'
        dvd = DataValidation(type='list', formula1=f'="{",".join(DIAS)}"', allow_blank=True)
        wm.add_data_validation(dvd); dvd.add(f'A{prim}:A{ult}')
        dvm = DataValidation(type='list', formula1='=marcas', allow_blank=True)
        wm.add_data_validation(dvm); dvm.add(f'E{prim}:I{ult}')
        for cod, cor, txt in [('P', OK_C, OK), ('F', 'FBE4E4', 'A33A3A'), ('FJ', ALERTA_C, ALERTA),
                              ('REP', MARCA_C, MARCA), ('LIC', CHEIO, GRAFITE)]:
            wm.conditional_formatting.add(f'E{prim}:I{ult}',
                CellIsRule(operator='equal', formula=[f'"{cod}"'], fill=fundo(cor), font=fonte(10, True, txt)))

    # --------------------------------------------------------- AutoFluxos
    wf = wb.create_sheet('AutoFluxos'); moldura(wf); impressao(wf, False)
    titulo(wf, 'A1', 'AutoFluxos')
    subtitulo(wf, 'A2', 'Saída lida pelo robô do WhatsApp. Cada linha traz os horários com vaga naquele dia. Não edite esta aba.')
    wf.merge_cells('A2:D2'); wf.row_dimensions[2].height = 28
    for col, w in zip('ABCD', [14, 50, 10, 22]): wf.column_dimensions[col].width = w
    p = cabecalho(wf, 4, ['Dia', 'Horários com vaga', 'Quantos', 'Intervalo nomeado'], None)
    for i, dia in enumerate(DIAS):
        r = p + i
        wf.cell(r, 1, dia).font = fonte(10, True)
        col = get_column_letter(9 + i)
        wf.cell(r, 5, f'=Grade!${col}${G2}')
        wf.cell(r, 2, f'=IF($E{r}="","",LEFT($E{r},LEN($E{r})-1))').font = fonte(10)
        wf.cell(r, 3, f'=IF($B{r}="",0,LEN($B{r})-LEN(SUBSTITUTE($B{r},";",""))+1)').alignment = CENTRO
        wf.cell(r, 4, SEM_ACENTO[i]).font = fonte(10, False, GRAFITE)
        for c in range(1, 5): wf.cell(r, c).border = borda()
    zebra(wf, p, p + 6, 4)
    wf.column_dimensions['E'].hidden = True
    wf.print_area = f'A1:D{p + 6}'
    for i, nome in enumerate(SEM_ACENTO):
        wb.defined_names.add(DefinedName(nome, attr_text=f'AutoFluxos!$B${p+i}'))

    # Cor de aba. Com 17 abas, achar o mês certo é o gesto mais repetido do
    # arquivo — e cor resolve isso antes de a pessoa ler o nome.
    cores = {'Início': MARCA, 'Config': GRAFITE, 'Grade': MARCA,
             'Alunos': MARCA, 'AutoFluxos': 'A9B4C4'}
    for nome, cor in cores.items(): wb[nome].sheet_properties.tabColor = cor
    for i, mes in enumerate(MESES, start=1):
        # Meses passados em tom apagado, o resto no tom da marca clara.
        wb[f'{i:02d} {mes[:3]}'].sheet_properties.tabColor = 'C9DDE1'

    wb.save(saida)
