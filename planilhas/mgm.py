# -*- coding: utf-8 -*-
"""Planilha 1 — os dados atuais do MGM Pilates, organizados.

Mesma informação da lista de turma dele, em formato que uma pessoa lê e um
programa também. Uma linha por turma e uma linha por aluno, em vez de blocos
desenhados — é o que permite acrescentar professor e horário sem quebrar nada.
"""
import openpyxl, unicodedata, re
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.defined_name import DefinedName
from estilo import *

DIAS_ORD = {'SEGUNDA':1,'TERÇA':2,'QUARTA':3,'QUINTA':4,'SEXTA':5,'SÁBADO':6}
BONITO = {'SEGUNDA':'Segunda','TERÇA':'Terça','QUARTA':'Quarta','QUINTA':'Quinta','SEXTA':'Sexta','SÁBADO':'Sábado'}
SEM_ACENTO = {'Segunda':'segunda','Terça':'terca','Quarta':'quarta','Quinta':'quinta','Sexta':'sexta','Sábado':'sabado'}

def norm_prof(p):
    if not p: return ''
    p = ' '.join(w.capitalize() for w in p.strip().split())
    troca = {'Marcia':'Márcia','Nathalia':'Nathália','Persio':'Pérsio'}
    return troca.get(p, p)

def norm_hora(h):
    m = re.match(r'(\d{1,2})\s*h\s*(\d{0,2})', str(h).strip(), re.I)
    if not m: return str(h).strip()
    return f'{int(m.group(1)):02d}:{(m.group(2) or "00").zfill(2)}'

def norm_fone(f):
    d = re.sub(r'\D', '', str(f or ''))
    if not d: return ''
    if len(d) == 9: d = '11' + d          # o formato dominante não traz DDD
    if len(d) == 8: d = '119' + d
    if len(d) == 10: d = d[:2] + '9' + d[2:]
    return f'({d[:2]}) {d[2:7]}-{d[7:11]}' if len(d) == 11 else str(f).strip()

def anotacao(nome):
    """As anotações que hoje moram dentro da célula do nome."""
    achados = re.findall(r'\(([^)]*)\)', nome)
    resto = re.sub(r'\([^)]*\)', '', nome)
    m = re.search(r'-\s*([A-ZÀ-Ú][A-ZÀ-Ú\s]{2,})$', resto.strip())
    if m: achados.append(m.group(1).strip()); resto = resto[:m.start()]
    return ' '.join(w.capitalize() for w in ' · '.join(achados).split()) if achados else '', ' '.join(resto.split())

def monta(turmas, saida):
    wb = openpyxl.Workbook()

    # ---------------------------------------------------------------- Início
    ws = wb.active; ws.title = 'Início'; moldura(ws)
    ws.column_dimensions['A'].width = 3
    for col, w in zip('BCDEFG', [22, 16, 16, 16, 16, 30]):
        ws.column_dimensions[col].width = w
    titulo(ws, 'B2', 'MGM Pilates', 20)
    subtitulo(ws, 'B3', 'Turmas e alunos — dados de agosto/26, reorganizados a partir da lista de turma atual.')
    ws.merge_cells('B3:G3'); ws.row_dimensions[3].height = 30

    total_t = len(turmas)
    total_v = sum(t['capacidade'] for t in turmas)
    total_m = sum(len([a for a in t['alunos'] if a['na_vaga']]) for t in turmas)
    total_e = sum(len([a for a in t['alunos'] if not a['na_vaga']]) for t in turmas)

    cartoes = [('Turmas por semana', total_t), ('Vagas', total_v),
               ('Matriculadas', total_m), ('Vagas livres', total_v - total_m),
               ('Encaixes', total_e)]
    for i, (rot, val) in enumerate(cartoes):
        col = 2 + i
        c = ws.cell(6, col, val); c.font = fonte(22, True, MARCA); c.alignment = CENTRO
        r = ws.cell(7, col, rot); r.font = fonte(9, False, GRAFITE); r.alignment = CENTRO
        for lin in (6, 7):
            ws.cell(lin, col).fill = fundo(MARCA_C)
    ws.row_dimensions[6].height = 34; ws.row_dimensions[7].height = 20

    c = ws.cell(9, 2, f'Ocupação: {round(100*total_m/total_v)}% das vagas')
    c.font = fonte(11, True, ALERTA)

    guia = [
        ('O que mudou', 'A informação é a mesma. O formato é que passou a ser uma linha por turma e uma linha por aluno, em vez de blocos desenhados.'),
        ('Por que isso importa', 'Assim dá para acrescentar horário, professor ou aluno sem redesenhar nada — e um programa consegue ler.'),
        ('Turmas', 'Cada linha é um horário fixo da semana. "Livres" se calcula sozinho a partir de quem está matriculado.'),
        ('Alunos', 'Cada linha é uma pessoa numa turma. Quem é da vaga fixa aparece como Matrícula; quem é encaixe aparece como Encaixe.'),
        ('AutoFluxos', 'Aba de saída, lida pelo robô do WhatsApp. Não precisa mexer nela.'),
    ]
    linha = 12
    for tit, txt in guia:
        c = ws.cell(linha, 2, tit); c.font = fonte(10, True, TINTA)
        d = ws.cell(linha, 3, txt); d.font = fonte(10, False, GRAFITE); d.alignment = QUEBRA
        ws.merge_cells(start_row=linha, start_column=3, end_row=linha, end_column=7)
        ws.row_dimensions[linha].height = 30
        linha += 1

    # ---------------------------------------------------------------- Turmas
    wt = wb.create_sheet('Turmas'); moldura(wt)
    titulo(wt, 'A1', 'Turmas'); wt.row_dimensions[1].height = 28
    cols = ['Dia', 'Hora', 'Turma', 'Professor', 'Capacidade', 'Matriculadas', 'Livres', 'Situação']
    prim = cabecalho(wt, 3, cols, [12, 9, 10, 14, 12, 13, 9, 14])
    turmas_ord = sorted(turmas, key=lambda t: (DIAS_ORD.get(t['dia'], 9), t['hora_norm']))
    r = prim
    for t in turmas_ord:
        chave = f"{BONITO[t['dia']]} {t['hora_norm']}"
        wt.cell(r, 1, BONITO[t['dia']]).alignment = ESQ
        wt.cell(r, 2, t['hora_norm']).alignment = CENTRO
        wt.cell(r, 3, t['codigo']).alignment = CENTRO
        wt.cell(r, 4, t['prof_norm']).alignment = ESQ
        wt.cell(r, 5, t['capacidade']).alignment = CENTRO
        wt.cell(r, 6, f'=COUNTIFS(Alunos!$G:$G,$A{r}&" "&$B{r},Alunos!$H:$H,"Matrícula")').alignment = CENTRO
        wt.cell(r, 7, f'=MAX(0,$E{r}-$F{r})').alignment = CENTRO
        wt.cell(r, 8, 'Fechada' if t['fechado'] else f'=IF($G{r}=0,"Cheia","Tem vaga")').alignment = CENTRO
        for c in range(1, 9):
            wt.cell(r, c).font = fonte(10)
            wt.cell(r, c).border = borda()
        r += 1
    ult = r - 1
    zebra(wt, prim, ult, 8)
    wt.freeze_panes = f'A{prim}'
    wt.auto_filter.ref = f'A3:H{ult}'
    wt.conditional_formatting.add(f'H{prim}:H{ult}',
        CellIsRule(operator='equal', formula=['"Tem vaga"'], fill=fundo(OK_C), font=fonte(10, True, OK)))
    wt.conditional_formatting.add(f'H{prim}:H{ult}',
        CellIsRule(operator='equal', formula=['"Cheia"'], fill=fundo(CHEIO), font=fonte(10, False, GRAFITE)))
    wt.conditional_formatting.add(f'H{prim}:H{ult}',
        CellIsRule(operator='equal', formula=['"Fechada"'], fill=fundo(ALERTA_C), font=fonte(10, True, ALERTA)))

    # ---------------------------------------------------------------- Alunos
    wa = wb.create_sheet('Alunos'); moldura(wa)
    titulo(wa, 'A1', 'Alunos'); wa.row_dimensions[1].height = 28
    cols = ['Matrícula', 'Nome', 'Telefone', 'Venc. do plano', 'Próxima avaliação', 'Anotações', 'Turma', 'Vínculo']
    prim_a = cabecalho(wa, 3, cols, [11, 34, 16, 14, 17, 20, 18, 11])
    r = prim_a
    for t in turmas_ord:
        chave = f"{BONITO[t['dia']]} {t['hora_norm']}"
        for a in t['alunos']:
            anota, nome = anotacao(a['nome'])
            wa.cell(r, 1, a['matricula']).alignment = CENTRO
            wa.cell(r, 2, nome.title()).alignment = ESQ
            wa.cell(r, 3, norm_fone(a['fone'])).alignment = CENTRO
            cv = wa.cell(r, 4, a['venc']); cv.alignment = CENTRO; cv.number_format = 'dd/mm/yyyy'
            cp = wa.cell(r, 5, a['aval']); cp.alignment = CENTRO; cp.number_format = 'dd/mm/yyyy'
            wa.cell(r, 6, anota).alignment = ESQ
            wa.cell(r, 7, chave).alignment = ESQ
            wa.cell(r, 8, 'Matrícula' if a['na_vaga'] else 'Encaixe').alignment = CENTRO
            for c in range(1, 9):
                wa.cell(r, c).font = fonte(10)
                wa.cell(r, c).border = borda()
            r += 1
    ult_a = r - 1
    zebra(wa, prim_a, ult_a, 8)
    wa.freeze_panes = f'A{prim_a}'
    wa.auto_filter.ref = f'A3:H{ult_a}'
    dv = DataValidation(type='list', formula1='"Matrícula,Encaixe"', allow_blank=False)
    wa.add_data_validation(dv); dv.add(f'H{prim_a}:H{ult_a+200}')
    wa.conditional_formatting.add(f'H{prim_a}:H{ult_a}',
        CellIsRule(operator='equal', formula=['"Encaixe"'], fill=fundo(ALERTA_C), font=fonte(10, False, ALERTA)))

    # ------------------------------------------------------------ AutoFluxos
    wf = wb.create_sheet('AutoFluxos'); moldura(wf)
    titulo(wf, 'A1', 'AutoFluxos')
    subtitulo(wf, 'A2', 'Aba de saída, lida pelo robô do WhatsApp. Cada célula traz os horários com vaga naquele dia. Não precisa editar.')
    wf.merge_cells('A2:D2'); wf.row_dimensions[2].height = 30
    for col, w in zip('ABCD', [14, 46, 12, 40]): wf.column_dimensions[col].width = w
    p = cabecalho(wf, 4, ['Dia', 'Horários com vaga', 'Quantos', 'Intervalo nomeado'], None)
    r = p
    for dia in ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado']:
        wf.cell(r, 1, dia).font = fonte(10, True)
        f = (f'=IFERROR(TEXTJOIN(";",1,FILTER(Turmas!$B${prim}:$B${ult},'
             f'Turmas!$A${prim}:$A${ult}=$A{r},Turmas!$G${prim}:$G${ult}>0,'
             f'Turmas!$H${prim}:$H${ult}<>"Fechada")),"")')
        c = wf.cell(r, 2, f); c.font = fonte(10); c.alignment = ESQ
        wf.cell(r, 3, f'=IF($B{r}="",0,LEN($B{r})-LEN(SUBSTITUTE($B{r},";",""))+1)').alignment = CENTRO
        wf.cell(r, 4, SEM_ACENTO[dia]).font = fonte(10, False, GRAFITE)
        for c in range(1, 5): wf.cell(r, c).border = borda()
        r += 1
    zebra(wf, p, r-1, 4)
    for i, dia in enumerate(['segunda','terca','quarta','quinta','sexta','sabado']):
        wb.defined_names.add(DefinedName(dia, attr_text=f"AutoFluxos!$B${p+i}"))

    wb.save(saida)
    return total_t, total_m, ult_a - prim_a + 1
