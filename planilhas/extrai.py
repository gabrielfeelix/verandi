import openpyxl, json, re, unicodedata

DIAS = ['SEGUNDA','TERÇA','QUARTA','QUINTA','SEXTA','SÁBADO']

def limpo(v):
    return '' if v is None else str(v).strip()

def extrai(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    turmas, alunos = [], []
    for ws in wb.worksheets:
        if ws.max_row < 5: continue
        dia = ws.title
        heads = [r for r in range(1, ws.max_row+1) if limpo(ws.cell(r,1).value) == 'Vagas']
        for i, h in enumerate(heads):
            fim = heads[i+1]-1 if i+1 < len(heads) else ws.max_row
            prof = ''
            for rr in range(max(1,h-3), h):
                v = limpo(ws.cell(rr,1).value)
                if v.startswith('Prof'): prof = v.replace('Prof.','').strip()
            hora, cap, cod = '', 0, ''
            linhas = []
            for r in range(h+1, fim+1):
                a = ws.cell(r,1).value
                b, c = limpo(ws.cell(r,2).value), limpo(ws.cell(r,3).value)
                if b and not hora: hora = b
                if c and not cod: cod = c
                nome = limpo(ws.cell(r,5).value)
                if isinstance(a, int): cap += 1
                if nome and 'N O M E' not in nome:
                    linhas.append({
                        'nome': nome,
                        'matricula': limpo(ws.cell(r,4).value),
                        'fone': limpo(ws.cell(r,6).value),
                        'venc': ws.cell(r,7).value,
                        'aval': ws.cell(r,8).value,
                        'na_vaga': isinstance(a, int),
                    })
            if not cap and not hora: continue
            # horário fechado é escrito na célula do nome
            fechado = any('fech' in l['nome'].lower() for l in linhas)
            linhas = [l for l in linhas if 'fech' not in l['nome'].lower()]
            turmas.append({'dia': dia, 'hora': hora, 'prof': prof, 'codigo': cod,
                           'capacidade': cap, 'fechado': fechado,
                           'alunos': linhas})
            for l in linhas:
                alunos.append({**l, 'dia': dia, 'hora': hora, 'prof': prof})
    return turmas, alunos

t, a = extrai('/home/gabfelix/dev/4yu-apps/autofluxos/LISTA DE TURMA E PRESENÇA - AGOSTO 26.xlsx'.replace('/home','/home'))
print('ERRO ESPERADO') if False else None
